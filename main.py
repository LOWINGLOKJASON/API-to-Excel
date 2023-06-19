# ****************************************************************
# Name: Wing Lok LO
# Link: https://replit.com/join/utnxrxfqyp-lowinglokjason
# ****************************************************************

# Import packages
import requests
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, ScatterChart, PieChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart.label import DataLabelList
import pandas as pd

# Get data from weather API
api_key = "fe8864f0805d25db6c7d08959d89dd60"
cities = ["New York", "London", "Paris"]

# Create a virtual workbook
workbook = Workbook()
sheet1 = workbook.active
sheet1.title = "Weather Data"

# Write header row
sheet1["A1"] = "City"
sheet1["B1"] = "Temperature (Celsius)"
sheet1["C1"] = "Humidity (%)"
sheet1["D1"] = "Wind Speed (m/s)"
sheet1["E1"] = "Cloudiness (%)"

# Retrieve weather data for each city
for row, city in enumerate(cities, start=2):
  # Prepare the API request URL for each city
  url = f"http://api.openweathermap.org/data/2.5/weather?appid={api_key}&q={city}"
  
  # Send the API request
  response = requests.get(url)
  data = response.json()
  
  # Retrieve weather data for the cities
  # Convert temperature from Kelvin to Celsius
  temperature = data["main"]["temp"] - 273.15  
  humidity = data["main"]["humidity"]
  wind_speed = data["wind"]["speed"]
  cloudiness = data["clouds"]["all"]

  # Deploy cells 
  sheet1.cell(row=row, column=1, value=city)
  sheet1.cell(row=row, column=2, value=temperature)
  sheet1.cell(row=row, column=3, value=humidity)
  sheet1.cell(row=row, column=4, value=wind_speed)
  sheet1.cell(row=row, column=5, value=cloudiness)

# Create additional sheets for charts
chart_sheet = workbook.create_sheet(title="Charts")

# Create Bar Chart
bar_chart = BarChart()
bar_chart.title = "Temperature (Celsius)"
bar_chart.x_axis.title = "City"
bar_chart.y_axis.title = "Temperature"
data = Reference(sheet1, min_col=2, min_row=1, max_row=len(cities) + 1, max_col=2)
categories = Reference(sheet1, min_col=1, min_row=2, max_row=len(cities) + 1)
bar_chart.add_data(data, titles_from_data=True)
bar_chart.set_categories(categories)
chart_sheet.add_chart(bar_chart, "A1")

# Create Line Chart
line_chart = LineChart()
line_chart.title = "Humidity (%)"
line_chart.x_axis.title = "City"
line_chart.y_axis.title = "Humidity"
data = Reference(sheet1, min_col=3, min_row=1, max_row=len(cities) + 1, max_col=3)
categories = Reference(sheet1, min_col=1, min_row=2, max_row=len(cities) + 1)
line_chart.add_data(data, titles_from_data=True)
line_chart.set_categories(categories)
chart_sheet.add_chart(line_chart, "M1")

# Create Scatter Chart
scatter_chart = ScatterChart()
scatter_chart.title = "Wind Speed (m/s) vs. Cloudiness (%)"
scatter_chart.x_axis.title = "Wind Speed"
scatter_chart.y_axis.title = "Cloudiness"
x_values = Reference(sheet1, min_col=4, min_row=1, max_row=len(cities) + 1, max_col=4)
y_values = Reference(sheet1, min_col=5, min_row=1, max_row=len(cities) + 1, max_col=5)
scatter_chart.add_data(x_values, y_values)
chart_sheet.add_chart(scatter_chart, "A16")

# Create Pie Chart
pie_chart = PieChart()
pie_chart.title = "Cloudiness (%)"
data = Reference(sheet1, min_col=5, min_row=1, max_row=len(cities) + 1, max_col=5)
categories = Reference(sheet1, min_col=1, min_row=2, max_row=len(cities) + 1)
pie_chart.add_data(data, titles_from_data=True)
pie_chart.set_categories(categories)
chart_sheet.add_chart(pie_chart, "M16")

# Save the workbook as an Excel file
workbook.save("weather_data.xlsx")
print("Done")
